import streamlit as st
import subprocess
import os
import glob
# import pyautogui as ag
# import pyperclip
from time import sleep
# import win32print
# import win32api
import openpyxl
from typing import List, Dict, Tuple


path_ueno = r"\\dwhnas1\DWH1\u_ueno\c_cuttingdata_auto"
path_cuttingfolder = r"\\dwhnas1\DWH1\泉州電業\泉州_切断資料_自動化用"
path_label_template_folder = r"\\dwhnas1\DWH1\u_ueno\c_cuttingdata_auto\s切断資料_雛形_保管場所"

st.title("切断資料自動作成")

# 選択欄から切断資料を指定
sep = "\\" if os.name == "nt" else "/"
file_paths = sorted(glob.glob(os.path.join(path_cuttingfolder, "*.xlsx")))

filename_lst_duplication = [""]
for file_path in file_paths:
    filename = file_path.split(sep)[-1].replace(f".xlsx", "")
    filename_lst_duplication.append(filename)
filename_list = list(dict.fromkeys(filename_lst_duplication))
filename = st.selectbox("切断資料ファイル名（選択）", filename_list)

# 入力欄から切断資料を指定
if filename == "":
    filename = st.text_input("切断資料ファイル名（入力）")

path = os.path.join(path_cuttingfolder, f"{filename}.xlsx")

"""
追加で印刷するシートの選択
"""

wb = openpyxl.load_workbook(path)
# 切断資料のシートを取得
sheet_main = wb[filename]
# 全シートを取得
sheet_all = wb.sheetnames

col = st.columns(len(sheet_all))

sheet_active_flag = []
for i, sheet in enumerate(sheet_all):
    if i == 0:
        sheet_active_flag.append(True)
    else:
        sheet_ = col[i].checkbox(label=sheet)
        sheet_active_flag.append(sheet_)


# 雛形選択
path_label_template = sorted(
    glob.glob(os.path.join(path_label_template_folder, "*.jlb"))
)
label_template_lst_duplication = []
for path_label in path_label_template:
    templatename = path_label.split(sep)[-1].replace(".jlb", "")
    label_template_lst_duplication.append(templatename)
label_template_lst = list(dict.fromkeys(label_template_lst_duplication))

label_template = st.radio("ラベル雛形", label_template_lst)
label_template = os.path.join(
    r"\\dwhnas1\DWH1\u_ueno\c_cuttingdata_auto\s切断資料_雛形_保管場所", label_template
)


def sheet_activate(sheet_all: List[str], path: str, sheet_activate_flag: List[bool]):
    wb = openpyxl.load_workbook(path)
    for sheet in sheet_all:
        ws = wb[sheet]
        ws.sheet_view.tabSelected = False
    wb.save(path)
    wb.close()

    wb = openpyxl.load_workbook(path)

    for i, active in enumerate(sheet_activate_flag):
        if active:
            ws = wb.worksheets[i]
            wb.active = i
            ws.sheet_view.tabSelected = True
        elif not active:
            pass
    wb.save(path)


def sheet_1(sheet_all: List[str], filename: str):
    wb = openpyxl.load_workbook(path)
    for sheet in sheet_all:
        ws = wb[sheet]
        if sheet == filename:
            ws.sheet_view.tabSelected = True
        else:
            ws.sheet_view.tabSelected = False
    wb.save(path)
    wb.close()


# サイドバー
st.sidebar.header("編集画面")
edit_form = st.sidebar.form("edit_form")
disconnection_date = edit_form.date_input("切断期日")
order_number = edit_form.number_input("受注番号", 0, 1000000, 0)
number_of_order = edit_form.number_input("受注数", 0, 10000, 0)
deadline = edit_form.date_input("納期")
push_button = edit_form.form_submit_button()


# Excelにて切断資料を印刷する
def PrintOut():
    win32api.ShellExecute(
        0, "print", path, "/c:" "%s" % win32print.GetDefaultPrinter(), ".", 0
    )


# ラベル編集ソフトを用いて，編集した切断資料を読み込んで印刷する
def label_main(excel_path: str, label_template: str):
    # ラベルソフトで正常に切断資料を読み込むために，切断資料を開いて保存したのち閉じる
    subprocess.Popen(["start", excel_path], shell=True)
    sleep(5)
    ag.hotkey("ctrl", "s")
    sleep(1)
    ag.hotkey("alt", "f4")
    sleep(3)

    # ラベル編集ソフトを起動する
    ag.press("win")
    sleep(1)
    pyperclip.copy("ラベルマイティ16 プレミアム")
    ag.hotkey("ctrl", "v")
    sleep(2)
    ag.press("Enter")
    sleep(3)

    # 切断資料の雛形を読み込む
    ag.press("tab", presses=3, interval=0.5)

    ag.press("Enter")
    sleep(1.5)
    ag.hotkey("alt", "n")
    pyperclip.copy(label_template + ".jlb")
    ag.hotkey("ctrl", "v")
    ag.press("Enter")
    sleep(2)

    # 「データ差込・連番の設定」を選択する
    p = ag.locateOnScreen((r"C:\label_picture\data_setting_ueno.png"), confidence=0.8)
    x, y = ag.center(p)
    ag.click(x, y)
    sleep(2)

    # 編集した切断資料を読み込む
    q = ag.locateOnScreen((r"C:\label_picture\load_data_ueno.png"), confidence=0.8)
    x, y = ag.center(q)
    ag.click(x, y)
    ag.hotkey("alt", "t")
    sleep(1)
    ag.press("down", presses=10)
    sleep(1)
    ag.hotkey("alt", "n")
    sleep(1)
    pyperclip.copy(excel_path)
    sleep(1)
    ag.hotkey("ctrl", "v")
    sleep(2)
    ag.press("Enter")
    sleep(2)
    ag.press("up", presses=10)
    sleep(1)
    ag.press("r")
    sleep(1)
    ag.press("t")
    sleep(1)
    ag.press("Enter")
    sleep(1)
    ag.press("Enter")
    sleep(1)

    # プリンタを設定する
    ag.hotkey("ctrl", "p")
    sleep(1)
    ag.press("tab")
    sleep(1)
    ag.press("Enter")
    sleep(1)
    ag.hotkey("hanja")
    sleep(1)
    ag.press("s")
    sleep(1)
    ag.hotkey("alt", "z")
    sleep(1)
    ag.press("up", presses=10)
    sleep(1)
    ag.press("Enter")
    sleep(1)


def label_print():
    print("a")
    print(ag.getWindowsWithTitle("ラベルマイティプレミアム"))
    labelprinter_window = ag.getWindowsWithTitle("ラベルマイティプレミアム")[0]
    labelprinter_window.activate()

    # 切断資料を印刷する
    ag.hotkey("shift", "tab")
    sleep(1)
    ag.press("Enter")

    """
    sleep(10)
    
    # ラベル編集ソフトを閉じる
    ag.hotkey("alt", "f4")
    sleep(0.5)
    ag.press("n")
    st.warning("ラベルを印刷しました！")
    """


# Streamlitでsubmitした後の処理
if push_button:
    # デフォルトのプリンタをSATOからFUJIFILMに変更する
    win32print.SetDefaultPrinterW("FUJIFILM Apeos C2570")

    # 1-4行目を表示する
    wb = openpyxl.load_workbook(path)
    sheet_main = wb[filename]
    for i in range(1, 5):
        sheet_main.row_dimensions[i].hidden = False
    wb.save(path)
    wb.close()

    # 切断期日，受注番号，受注数，納期を切断資料に反映する
    sheet_main["O1"] = disconnection_date.strftime("%Y/%m/%d")
    sheet_main["O2"].value = order_number
    sheet_main["O3"].value = number_of_order
    sheet_main["O4"] = deadline.strftime("%Y/%m/%d")
    wb.save(path)
    wb.close()

    # 1-4行目を表示させたまま切断資料を印刷するために．切断資料を開いて保存したのち閉じる
    subprocess.Popen(["start", path], shell=True)
    sleep(5)
    ag.hotkey("ctrl", "s")
    sleep(1)
    ag.hotkey("alt", "f4")
    sleep(5)

    # 選択したシートをアクティブ化する
    sheet_activate(sheet_all, path, sheet_active_flag)

    # 切断資料を印刷する
    # PrintOut()
    sleep(5)
    st.warning("切断資料を印刷しました！")

    # 1-4行目を非表示にする
    wb = openpyxl.load_workbook(path)
    sheet_main = wb[filename]
    for i in range(1, 5):
        sheet_main.row_dimensions[i].hidden = True
    wb.save(path)
    wb.close()

    # 切断資料のみアクティブ化する
    sheet_1(sheet_all, filename)

    # ラベルプリンタを印刷する
    label_main(path, label_template)
